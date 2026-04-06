from __future__ import annotations

import argparse
import shutil
from datetime import datetime
from pathlib import Path

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook


def _ts() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _safe_name(name: str) -> str:
    return "".join(ch if ch.isalnum() else "_" for ch in name).strip("_").lower()


def _sheet_preview_to_png(workbook_path: Path, sheet_name: str, out_path: Path, max_rows: int = 20, max_cols: int = 10) -> None:
    df = pd.read_excel(workbook_path, sheet_name=sheet_name, engine="openpyxl")
    if df is None:
        df = pd.DataFrame()
    df = df.head(max_rows)
    if len(df.columns) > max_cols:
        df = df.iloc[:, :max_cols]
    if df.empty:
        df = pd.DataFrame({"info": [f"Sheet '{sheet_name}' has no tabular preview rows."]})

    fig, ax = plt.subplots(figsize=(14, 6))
    ax.axis("off")
    tab = ax.table(
        cellText=df.fillna("").astype(str).values,
        colLabels=[str(c) for c in df.columns],
        loc="center",
    )
    tab.auto_set_font_size(False)
    tab.set_fontsize(8)
    tab.scale(1.05, 1.35)
    ax.set_title(f"Sheet Preview: {sheet_name} (top {len(df)} rows)", fontsize=12, pad=12)
    fig.tight_layout()
    fig.savefig(out_path, dpi=170)
    plt.close(fig)


def generate_artifacts(
    workbook_path: Path,
    source_image: Path,
    charts_dir: Path,
    out_dir: Path,
) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    screenshots_dir = out_dir / "screenshots"
    screenshots_dir.mkdir(parents=True, exist_ok=True)

    run_log = out_dir / "run_log.txt"
    lines: list[str] = []

    lines.append(f"[{_ts()}] START: Extension run artifact generation")
    lines.append(f"[{_ts()}] Workbook: {workbook_path}")
    lines.append(f"[{_ts()}] Source image: {source_image}")
    lines.append(f"[{_ts()}] Charts dir: {charts_dir}")

    # Stage 1: input image
    img_copy = screenshots_dir / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_stage01_input_image.png"
    shutil.copy2(source_image, img_copy)
    lines.append(f"[{_ts()}] Stage 01 (image extract input) screenshot: {img_copy}")

    # Stage 2+: sheet previews
    wb = load_workbook(workbook_path, data_only=True)
    ordered_sheets = wb.sheetnames
    wb.close()

    stage_map = [
        ("Original Data", "stage02_original_data"),
        ("Image Extracted Data", "stage03_image_extracted"),
        ("Dataset_A", "stage04_dataset_a"),
        ("Dataset_B", "stage05_dataset_b"),
        ("Individual Analysis - Original", "stage06_individual_analysis_original"),
        ("Individual Analysis - Image Data", "stage07_individual_analysis_image"),
        ("Fused Dataset", "stage08_fused_dataset"),
        ("Combined Analysis", "stage09_combined_analysis"),
        ("Final Comparison", "stage10_final_comparison"),
    ]

    for sheet_name, stage_label in stage_map:
        if sheet_name not in ordered_sheets:
            lines.append(f"[{_ts()}] WARN: Missing expected sheet: {sheet_name}")
            continue
        out_png = screenshots_dir / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{stage_label}.png"
        _sheet_preview_to_png(workbook_path, sheet_name, out_png)
        lines.append(f"[{_ts()}] {stage_label} screenshot: {out_png}")

    # Stage charts: copy chart PNG files with timestamped names
    if charts_dir.exists():
        chart_files = sorted(charts_dir.glob("*.png"))
        for idx, ch in enumerate(chart_files, start=1):
            dst = screenshots_dir / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_stage11_chart_{idx:02d}_{_safe_name(ch.stem)}.png"
            shutil.copy2(ch, dst)
            lines.append(f"[{_ts()}] stage11 chart screenshot: {dst}")
    else:
        lines.append(f"[{_ts()}] WARN: charts directory not found: {charts_dir}")

    lines.append(f"[{_ts()}] COMPLETE: Artifact generation done")
    run_log.write_text("\n".join(lines) + "\n", encoding="utf-8")

    # simple index
    index_md = out_dir / "index.md"
    md_lines = [
        "# Extension Run Log Artifacts",
        "",
        f"- Generated at: `{_ts()}`",
        f"- Workbook: `{workbook_path}`",
        f"- Source image: `{source_image}`",
        f"- Run log: `{run_log}`",
        "",
        "## Screenshot Files",
    ]
    for p in sorted(screenshots_dir.glob("*.png")):
        md_lines.append(f"- `{p.name}`")
    index_md.write_text("\n".join(md_lines) + "\n", encoding="utf-8")

    return out_dir


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate timestamped run log + screenshots for extension outputs.")
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--source-image", required=True)
    parser.add_argument("--charts-dir", required=True)
    parser.add_argument("--out-dir", required=False, default="")
    args = parser.parse_args()

    workbook = Path(args.workbook).expanduser().resolve()
    source_image = Path(args.source_image).expanduser().resolve()
    charts_dir = Path(args.charts_dir).expanduser().resolve()

    if args.out_dir:
        out_dir = Path(args.out_dir).expanduser().resolve()
    else:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_dir = workbook.parent / f"run_log_artifacts_{stamp}"

    out = generate_artifacts(workbook, source_image, charts_dir, out_dir)
    print(f"[OK] Artifacts generated at: {out}")


if __name__ == "__main__":
    main()
