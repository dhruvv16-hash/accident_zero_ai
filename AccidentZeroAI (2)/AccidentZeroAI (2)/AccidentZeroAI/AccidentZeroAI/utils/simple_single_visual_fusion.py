from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows


BASE_COLS = [
    "shift_hours",
    "overtime_hours",
    "worker_experience",
    "equipment_age",
    "maintenance_score",
    "temperature",
    "humidity",
    "inspection_score",
    "accident",
]


def _read_first_sheet(excel_path: Path) -> pd.DataFrame:
    sheets = pd.read_excel(excel_path, sheet_name=None, engine="openpyxl")
    for _, df in sheets.items():
        if df is not None and not df.empty:
            return df.copy()
    raise ValueError("No non-empty sheet found in input Excel.")


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [str(c).strip().lower().replace(" ", "_") for c in out.columns]
    for c in BASE_COLS:
        if c not in out.columns:
            out[c] = np.nan
    out = out[BASE_COLS]
    for c in out.columns:
        out[c] = pd.to_numeric(out[c], errors="coerce")
    return out


def _generate_image_tables_from_excel(df: pd.DataFrame, image_count: int, out_dir: Path) -> tuple[list[Path], pd.DataFrame]:
    out_dir.mkdir(parents=True, exist_ok=True)
    image_paths: list[Path] = []
    idx_groups = np.array_split(np.arange(len(df)), image_count)
    image_rows: list[pd.DataFrame] = []

    rng = np.random.default_rng(42)
    base = df.reset_index(drop=True)
    for i, idxs in enumerate(idx_groups, start=1):
        g = base.iloc[idxs].copy()
        if g.empty:
            continue
        g = g.copy().head(20)
        # simulate OCR/missing cells in image-derived data
        miss_mask = rng.random(g.shape) < 0.12
        g_missing = g.mask(miss_mask)
        image_rows.append(g_missing)

        fig, ax = plt.subplots(figsize=(12, 4))
        ax.axis("off")
        tbl = ax.table(
            cellText=g_missing.fillna("").astype(str).values,
            colLabels=g_missing.columns.tolist(),
            loc="center",
        )
        tbl.auto_set_font_size(False)
        tbl.set_fontsize(8)
        tbl.scale(1, 1.35)
        ax.set_title(f"Generated Image Table {i}", pad=8)
        img_path = out_dir / f"generated_image_table_{i}.png"
        fig.tight_layout()
        fig.savefig(img_path, dpi=170)
        plt.close(fig)
        image_paths.append(img_path)

    image_df = pd.concat(image_rows, ignore_index=True) if image_rows else pd.DataFrame(columns=df.columns)
    return image_paths, image_df


def _ai_impute_knn_like(df: pd.DataFrame) -> pd.DataFrame:
    """
    Lightweight nearest-neighbor imputation (AI-style) without external ML dependencies.
    """
    out = df.copy().reset_index(drop=True)
    num_cols = out.columns.tolist()

    # fallback medians
    med = out.median(numeric_only=True)
    for col in num_cols:
        if col not in med.index or pd.isna(med[col]):
            med[col] = 0.0

    arr = out[num_cols].to_numpy(dtype=float, copy=True)
    for i in range(arr.shape[0]):
        row = arr[i]
        miss = np.isnan(row)
        if not miss.any():
            continue

        avail = ~miss
        # candidate rows that have target col values and overlap on available cols
        distances = []
        for j in range(arr.shape[0]):
            if i == j:
                continue
            cand = arr[j]
            overlap = avail & ~np.isnan(cand)
            if overlap.sum() == 0:
                continue
            d = np.linalg.norm(row[overlap] - cand[overlap])
            distances.append((j, d))
        distances.sort(key=lambda x: x[1])
        neighbors = [idx for idx, _ in distances[:5]]

        for c_idx, is_missing in enumerate(miss):
            if not is_missing:
                continue
            vals = []
            for n in neighbors:
                v = arr[n, c_idx]
                if not np.isnan(v):
                    vals.append(v)
            if vals:
                arr[i, c_idx] = float(np.mean(vals))
            else:
                arr[i, c_idx] = float(med[num_cols[c_idx]])

    out[num_cols] = arr
    return out


def _bar_visual(df: pd.DataFrame, title: str, out_path: Path) -> Path:
    means = df.mean(numeric_only=True).sort_index()
    fig, ax = plt.subplots(figsize=(11, 5))
    ax.bar(means.index, means.values, color="#2e7d32")
    ax.set_title(title)
    ax.set_ylabel("Mean Value")
    ax.tick_params(axis="x", rotation=25)
    fig.tight_layout()
    fig.savefig(out_path, dpi=170)
    plt.close(fig)
    return out_path


def _final_bar_visual(excel_df: pd.DataFrame, image_df: pd.DataFrame, fused_df: pd.DataFrame, out_path: Path) -> Path:
    excel_means = excel_df.mean(numeric_only=True)
    image_means = image_df.mean(numeric_only=True)
    fused_means = fused_df.mean(numeric_only=True)

    cols = [c for c in excel_means.index if c in image_means.index and c in fused_means.index]
    x = np.arange(len(cols))
    w = 0.28

    fig, ax = plt.subplots(figsize=(13, 6))
    ax.bar(x - w, [excel_means[c] for c in cols], width=w, label="Excel")
    ax.bar(x, [image_means[c] for c in cols], width=w, label="Image")
    ax.bar(x + w, [fused_means[c] for c in cols], width=w, label="Fused")
    ax.set_xticks(x)
    ax.set_xticklabels(cols, rotation=25, ha="right")
    ax.set_title("Final Comparison (Single Visualization Type: Bar)")
    ax.set_ylabel("Mean Value")
    ax.legend()
    fig.tight_layout()
    fig.savefig(out_path, dpi=170)
    plt.close(fig)
    return out_path


def _write_df(ws, df: pd.DataFrame, title: str, start_row: int = 1) -> int:
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True)
    r0 = start_row + 1
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=r0):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
        if r_idx == r0:
            for c_idx in range(1, len(row) + 1):
                ws.cell(row=r_idx, column=c_idx).font = Font(bold=True)
    return r0 + len(df) + 2


def _embed_image(ws, img_path: Path, cell: str) -> None:
    if not img_path.exists():
        return
    img = XLImage(str(img_path))
    img.width = 700
    img.height = 330
    ws.add_image(img, cell)


def run_pipeline(input_excel: Path, output_excel: Path, image_count: int) -> dict[str, Path]:
    base_dir = output_excel.parent
    artifacts_dir = base_dir / f"simple_visual_artifacts_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    artifacts_dir.mkdir(parents=True, exist_ok=True)

    raw_excel = _read_first_sheet(input_excel)
    excel_df = _normalize_columns(raw_excel)

    image_dir = artifacts_dir / "generated_images"
    image_paths, image_raw_df = _generate_image_tables_from_excel(excel_df, image_count=image_count, out_dir=image_dir)
    image_imputed_df = _ai_impute_knn_like(image_raw_df) if not image_raw_df.empty else image_raw_df.copy()

    excel_tagged = excel_df.copy()
    excel_tagged["source"] = "excel"
    image_tagged = image_imputed_df.copy()
    image_tagged["source"] = "image"
    fused_df = pd.concat([excel_tagged, image_tagged], ignore_index=True)

    excel_pct = (len(excel_tagged) / len(fused_df) * 100.0) if len(fused_df) else 0.0
    image_pct = (len(image_tagged) / len(fused_df) * 100.0) if len(fused_df) else 0.0

    # Single visualization type only: bar charts
    vis_dir = artifacts_dir / "visualizations"
    vis_dir.mkdir(parents=True, exist_ok=True)
    excel_vis = _bar_visual(excel_df, "Excel Input Visualization (Bar)", vis_dir / "excel_bar.png")
    image_vis = _bar_visual(image_imputed_df, "Image Input Visualization (Bar)", vis_dir / "image_bar.png")
    final_vis = _final_bar_visual(excel_df, image_imputed_df, fused_df[BASE_COLS], vis_dir / "final_fused_bar.png")

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Original Data"
    r = _write_df(ws1, excel_df, "Original Excel Input (unchanged)")
    _embed_image(ws1, excel_vis, "A25")
    ws1.cell(row=r, column=1, value="Visualization type used: Bar chart only").font = Font(bold=True)

    ws2 = wb.create_sheet("Image Data")
    r2 = _write_df(ws2, image_raw_df, "Generated Image-based Data (with missing values)")
    r2 = _write_df(ws2, image_imputed_df, "AI-Imputed Image Data", start_row=r2)
    _embed_image(ws2, image_vis, "A30")
    ws2.cell(row=r2, column=1, value=f"Generated image tables count: {len(image_paths)}").font = Font(bold=True)

    ws3 = wb.create_sheet("Fused Data")
    r3 = _write_df(ws3, fused_df, "Fused Data (Excel + Image)")
    ws3.cell(row=r3, column=1, value="Contribution (%)").font = Font(bold=True)
    ws3.cell(row=r3 + 1, column=1, value="Excel contribution %")
    ws3.cell(row=r3 + 1, column=2, value=round(excel_pct, 2))
    ws3.cell(row=r3 + 2, column=1, value="Image contribution %")
    ws3.cell(row=r3 + 2, column=2, value=round(image_pct, 2))
    _embed_image(ws3, final_vis, "A30")

    ws4 = wb.create_sheet("Run Log")
    log_df = pd.DataFrame(
        [
            {"step": "input_excel_loaded", "status": "ok", "details": str(input_excel)},
            {"step": "image_tables_generated", "status": "ok", "details": str(len(image_paths))},
            {"step": "ai_imputation_done", "status": "ok", "details": "nearest-neighbor imputation"},
            {"step": "fusion_completed", "status": "ok", "details": f"excel={len(excel_tagged)}, image={len(image_tagged)}"},
            {"step": "final_contribution_pct", "status": "ok", "details": f"excel={excel_pct:.2f}%, image={image_pct:.2f}%"},
        ]
    )
    _write_df(ws4, log_df, "Simple Single-Visualization Pipeline Log")

    output_excel.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_excel)

    # write external log txt
    log_txt = artifacts_dir / "run_log.txt"
    log_txt.write_text(
        "\n".join(
            [
                f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] input_excel={input_excel}",
                f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] generated_images={len(image_paths)}",
                f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] excel_rows={len(excel_tagged)}",
                f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] image_rows={len(image_tagged)}",
                f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] contribution_excel_pct={excel_pct:.2f}",
                f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] contribution_image_pct={image_pct:.2f}",
                f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] output_excel={output_excel}",
                f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] visual_excel={excel_vis}",
                f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] visual_image={image_vis}",
                f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] visual_final={final_vis}",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    return {
        "output_excel": output_excel,
        "artifacts_dir": artifacts_dir,
        "excel_vis": excel_vis,
        "image_vis": image_vis,
        "final_vis": final_vis,
        "run_log": log_txt,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Simple fusion pipeline with one visualization type (bar chart only).")
    parser.add_argument("--input-excel", required=True, help="Input Excel path.")
    parser.add_argument("--output-excel", required=True, help="New output Excel path.")
    parser.add_argument("--image-count", type=int, default=6, help="How many generated image tables to create.")
    args = parser.parse_args()

    res = run_pipeline(
        input_excel=Path(args.input_excel).expanduser().resolve(),
        output_excel=Path(args.output_excel).expanduser().resolve(),
        image_count=max(2, args.image_count),
    )
    print(f"[OK] Output Excel: {res['output_excel']}")
    print(f"[OK] Artifacts dir: {res['artifacts_dir']}")
    print(f"[OK] Run log: {res['run_log']}")


if __name__ == "__main__":
    main()
