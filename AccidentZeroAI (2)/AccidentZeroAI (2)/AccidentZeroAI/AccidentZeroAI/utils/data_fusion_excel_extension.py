from __future__ import annotations

import argparse
import base64
import json
import os
import re
import shutil
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

try:
    import seaborn as sns
except Exception:  # pragma: no cover
    sns = None

try:
    import matplotlib.pyplot as plt
except Exception:  # pragma: no cover
    plt = None


BASE_COLS = [
    "shift_hours",
    "overtime_hours",
    "worker_experience",
    "equipment_age",
    "maintenance_score",
    "temperature",
    "humidity",
    "inspection_score",
    "accident",
]

COL_ALIASES = {
    "shift_hours": {"shifthours", "shift_hrs", "shift", "hours_shift"},
    "overtime_hours": {"overtime", "ot_hours", "extra_hours"},
    "worker_experience": {"experience", "worker_exp", "years_experience"},
    "equipment_age": {"machine_age", "asset_age"},
    "maintenance_score": {"maintenance", "maint_score"},
    "temperature": {"temp", "temp_c", "temperature_c"},
    "humidity": {"humid", "humidity_percent", "rh"},
    "inspection_score": {"inspection", "audit_score"},
    "accident": {"incident", "accident_flag", "target"},
}


@dataclass
class ImageExtractionResult:
    raw_table: pd.DataFrame
    mapped_table: pd.DataFrame
    detected_content: str
    extraction_method: str
    assumptions: list[str]
    corrections: list[str]


def _slug(s: str) -> str:
    s = str(s).strip().lower()
    s = re.sub(r"([a-z])([A-Z])", r"\1_\2", s)
    s = re.sub(r"[\s\-/]+", "_", s)
    s = re.sub(r"[^a-z0-9_]", "", s)
    s = re.sub(r"_+", "_", s)
    return s.strip("_")


def _autofit_cols(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(60, max(12, max_len + 2))


def _write_df(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1, title: str | None = None) -> int:
    row_ptr = start_row
    if title:
        ws.cell(row=row_ptr, column=start_col, value=title).font = Font(bold=True)
        row_ptr += 1
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=row_ptr):
        for c_idx, val in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=val)
        if r_idx == row_ptr:
            for c_idx in range(start_col, start_col + len(row)):
                ws.cell(row=r_idx, column=c_idx).font = Font(bold=True)
    return row_ptr + len(df.index) + 2


def _coerce_numeric(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for c in out.columns:
        out[c] = out[c].replace({"": np.nan, "NULL": np.nan, "null": np.nan, "-": np.nan})
    for c in out.columns:
        try:
            out[c] = pd.to_numeric(out[c], errors="raise")
        except Exception:
            # Keep original non-numeric columns as-is.
            pass
    return out


def _map_to_meaningful_columns(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    mapped = pd.DataFrame(index=df.index)
    assumptions: list[str] = []

    slug_map: dict[str, str] = {_slug(c): c for c in df.columns}
    reverse_alias: dict[str, str] = {}
    for canonical, aliases in COL_ALIASES.items():
        reverse_alias[canonical] = canonical
        for a in aliases:
            reverse_alias[_slug(a)] = canonical

    used_input_cols: set[str] = set()
    for canonical in BASE_COLS:
        found_col: str | None = None
        for s_col, original_col in slug_map.items():
            if s_col == canonical or reverse_alias.get(s_col) == canonical:
                found_col = original_col
                break
        if found_col is not None:
            mapped[canonical] = df[found_col]
            used_input_cols.add(found_col)
        else:
            mapped[canonical] = np.nan
            assumptions.append(f"Column '{canonical}' not found in image table; filled with NaN.")

    leftovers = [c for c in df.columns if c not in used_input_cols]
    if leftovers:
        assumptions.append(
            f"Unmapped image columns retained separately in raw extract: {', '.join(map(str, leftovers[:8]))}."
        )
    return _coerce_numeric(mapped), assumptions


def _api_key() -> str:
    return (os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY") or "").strip().strip("'\"")


def _try_gemini_table_extract(image_bytes: bytes, image_mime: str) -> ImageExtractionResult | None:
    key = _api_key()
    if not key:
        return None

    api_base = (os.environ.get("GEMINI_API_BASE") or "https://generativelanguage.googleapis.com/v1beta").rstrip("/")
    model = (os.environ.get("GEMINI_MODEL") or "gemini-2.5-flash").replace("models/", "").strip()
    url = f"{api_base}/models/{model}:generateContent"

    system = (
        "You are an OCR and table-structure extraction engine. "
        "Extract tabular/numbered information from the image into a structured JSON object only."
    )
    user = (
        "Detect if the image has table/chart/handwritten or printed numeric data. "
        "Return ONLY JSON with keys: detected_content, columns, rows, assumptions, corrections. "
        "rows must be array of objects. If chart-only, estimate readable datapoints and include them."
    )
    b64 = base64.standard_b64encode(image_bytes).decode("ascii")
    payload = {
        "systemInstruction": {"role": "system", "parts": [{"text": system}]},
        "contents": [
            {
                "role": "user",
                "parts": [
                    {"text": user},
                    {"inline_data": {"mime_type": image_mime, "data": b64}},
                ],
            }
        ],
        "generationConfig": {"temperature": 0.2, "maxOutputTokens": 4096},
    }

    resp = requests.post(url, params={"key": key}, json=payload, timeout=90)
    if resp.status_code >= 400:
        return None

    j = resp.json()
    text_out = ""
    for cand in j.get("candidates", []):
        for part in cand.get("content", {}).get("parts", []):
            if "text" in part:
                text_out += str(part["text"])
    text_out = text_out.strip()
    m = re.search(r"\{[\s\S]*\}", text_out)
    if not m:
        return None
    parsed = json.loads(m.group(0))

    rows = parsed.get("rows") or []
    if not isinstance(rows, list) or len(rows) == 0:
        return None
    raw_df = pd.DataFrame(rows)
    if raw_df.empty:
        return None

    mapped_df, map_assumptions = _map_to_meaningful_columns(raw_df)
    assumptions = [str(x) for x in parsed.get("assumptions") or []] + map_assumptions
    corrections = [str(x) for x in parsed.get("corrections") or []]
    return ImageExtractionResult(
        raw_table=_coerce_numeric(raw_df),
        mapped_table=mapped_df,
        detected_content=str(parsed.get("detected_content") or "table/chart"),
        extraction_method="gemini_vision_structured_ocr",
        assumptions=assumptions,
        corrections=corrections,
    )


def _try_pytesseract_extract(image_path: Path) -> ImageExtractionResult | None:
    try:
        import pytesseract
        from PIL import Image
    except Exception:
        return None

    text = pytesseract.image_to_string(Image.open(image_path))
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if len(lines) < 2:
        return None

    # Heuristic parsing: split by tabs or repeated spaces.
    parsed_rows: list[list[str]] = []
    for ln in lines:
        parts = re.split(r"\t+|\s{2,}", ln)
        parts = [p.strip() for p in parts if p.strip()]
        if len(parts) >= 2:
            parsed_rows.append(parts)
    if len(parsed_rows) < 2:
        return None

    width = max(len(r) for r in parsed_rows)
    padded = [r + [""] * (width - len(r)) for r in parsed_rows]
    header = padded[0]
    raw_df = pd.DataFrame(padded[1:], columns=header)
    mapped_df, assumptions = _map_to_meaningful_columns(raw_df)
    assumptions.append("OCR parser used whitespace/table-line heuristics; verify low-confidence rows.")
    return ImageExtractionResult(
        raw_table=_coerce_numeric(raw_df),
        mapped_table=mapped_df,
        detected_content="ocr_text_table",
        extraction_method="pytesseract_ocr",
        assumptions=assumptions,
        corrections=[],
    )


def extract_image_data(image_path: Path) -> ImageExtractionResult:
    if not image_path.exists():
        raise FileNotFoundError(f"Image file not found: {image_path}")
    image_bytes = image_path.read_bytes()
    ext = image_path.suffix.lower()
    mime = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".webp": "image/webp",
    }.get(ext, "image/jpeg")

    gem = _try_gemini_table_extract(image_bytes, mime)
    if gem is not None:
        return gem

    ocr = _try_pytesseract_extract(image_path)
    if ocr is not None:
        return ocr

    raise RuntimeError(
        "Image extraction failed. Configure GEMINI_API_KEY (recommended) or install pytesseract + Tesseract OCR."
    )


def clean_for_analysis(df: pd.DataFrame) -> pd.DataFrame:
    out = _coerce_numeric(df.copy())
    # Keep raw outside; this cleaned copy is only for analysis.
    for c in out.columns:
        if pd.api.types.is_numeric_dtype(out[c]):
            if out[c].isna().all():
                out[c] = 0.0
            else:
                out[c] = out[c].fillna(out[c].median())
    return out


def stats_table(df: pd.DataFrame, dataset_name: str) -> pd.DataFrame:
    ndf = df.select_dtypes(include=[np.number]).copy()
    if ndf.empty:
        return pd.DataFrame({"dataset": [dataset_name], "metric": ["no_numeric_columns"], "value": ["n/a"]})
    rows: list[dict[str, Any]] = []
    for c in ndf.columns:
        s = pd.to_numeric(ndf[c], errors="coerce")
        mode_vals = s.mode(dropna=True)
        rows.extend(
            [
                {"dataset": dataset_name, "column": c, "metric": "mean", "value": float(s.mean())},
                {"dataset": dataset_name, "column": c, "metric": "median", "value": float(s.median())},
                {
                    "dataset": dataset_name,
                    "column": c,
                    "metric": "mode",
                    "value": float(mode_vals.iloc[0]) if not mode_vals.empty else np.nan,
                },
                {"dataset": dataset_name, "column": c, "metric": "std", "value": float(s.std(ddof=1))},
                {"dataset": dataset_name, "column": c, "metric": "variance", "value": float(s.var(ddof=1))},
                {"dataset": dataset_name, "column": c, "metric": "min", "value": float(s.min())},
                {"dataset": dataset_name, "column": c, "metric": "max", "value": float(s.max())},
            ]
        )
    return pd.DataFrame(rows)


def _safe_numeric_cols(df: pd.DataFrame) -> list[str]:
    cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
    return cols[:8]


def _require_plotting() -> None:
    if plt is None:
        raise RuntimeError("matplotlib is required for chart generation. Install dependencies from requirements.txt.")


def generate_individual_charts(df: pd.DataFrame, dataset_name: str, out_dir: Path) -> tuple[list[Path], list[str]]:
    _require_plotting()
    out_dir.mkdir(parents=True, exist_ok=True)
    ndf = df.select_dtypes(include=[np.number]).copy()
    chart_paths: list[Path] = []
    interpretations: list[str] = []

    if ndf.empty:
        return chart_paths, [f"{dataset_name}: no numeric columns available for chart generation."]

    cols = _safe_numeric_cols(ndf)
    x = np.arange(1, len(ndf) + 1)

    # Line chart
    fig, ax = plt.subplots(figsize=(10, 5))
    for c in cols[:4]:
        ax.plot(x, ndf[c], label=c, linewidth=1.5)
    ax.set_title(f"{dataset_name} - Line Chart")
    ax.set_xlabel("Row Index")
    ax.set_ylabel("Value")
    ax.grid(alpha=0.3)
    ax.legend(loc="best", fontsize=8)
    p1 = out_dir / f"{dataset_name}_line.png"
    fig.tight_layout()
    fig.savefig(p1, dpi=130)
    plt.close(fig)
    chart_paths.append(p1)
    interpretations.append(
        f"Line chart ({dataset_name}): shows row-wise trends; larger fluctuations indicate higher variability in observed factors."
    )

    # Bar chart (means)
    means = ndf[cols].mean().sort_values(ascending=False)
    fig, ax = plt.subplots(figsize=(10, 5))
    ax.bar(means.index, means.values, color="#2d6a4f")
    ax.set_title(f"{dataset_name} - Bar Chart (Column Means)")
    ax.set_ylabel("Mean Value")
    ax.tick_params(axis="x", rotation=30)
    p2 = out_dir / f"{dataset_name}_bar.png"
    fig.tight_layout()
    fig.savefig(p2, dpi=130)
    plt.close(fig)
    chart_paths.append(p2)
    interpretations.append(
        f"Bar chart ({dataset_name}): compares average magnitude across variables and highlights dominant factors."
    )

    # Histogram
    hist_col = cols[0]
    fig, ax = plt.subplots(figsize=(10, 5))
    ax.hist(ndf[hist_col].dropna(), bins=12, color="#40916c", edgecolor="black")
    ax.set_title(f"{dataset_name} - Histogram ({hist_col})")
    ax.set_xlabel(hist_col)
    ax.set_ylabel("Frequency")
    p3 = out_dir / f"{dataset_name}_hist.png"
    fig.tight_layout()
    fig.savefig(p3, dpi=130)
    plt.close(fig)
    chart_paths.append(p3)
    interpretations.append(
        f"Histogram ({dataset_name}): illustrates distribution shape for {hist_col}, useful for skewness or concentration checks."
    )

    # Scatter
    if len(cols) >= 2:
        fig, ax = plt.subplots(figsize=(10, 5))
        ax.scatter(ndf[cols[0]], ndf[cols[1]], alpha=0.7, color="#1b4332")
        ax.set_title(f"{dataset_name} - Scatter Plot")
        ax.set_xlabel(cols[0])
        ax.set_ylabel(cols[1])
        p4 = out_dir / f"{dataset_name}_scatter.png"
        fig.tight_layout()
        fig.savefig(p4, dpi=130)
        plt.close(fig)
        chart_paths.append(p4)
        interpretations.append(
            f"Scatter plot ({dataset_name}): evaluates pairwise relation between {cols[0]} and {cols[1]} for visible linear/nonlinear association."
        )
    else:
        interpretations.append(f"Scatter plot ({dataset_name}): skipped because fewer than two numeric variables were available.")

    return chart_paths, interpretations


def _embed_images(ws, image_paths: list[Path], start_cell: str) -> None:
    start_col = re.match(r"[A-Z]+", start_cell).group(0)  # type: ignore[union-attr]
    start_row = int(re.match(r"[A-Z]+(\d+)", start_cell).group(1))  # type: ignore[union-attr]
    col_ord = ord(start_col)
    row = start_row
    col = col_ord
    for i, p in enumerate(image_paths):
        if not p.exists():
            continue
        img = XLImage(str(p))
        img.width = 460
        img.height = 250
        cell = f"{chr(col)}{row}"
        ws.add_image(img, cell)
        if i % 2 == 0:
            col = col_ord + 8
        else:
            col = col_ord
            row += 18


def split_datasets(original_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, str]:
    if "accident" in original_df.columns:
        a = original_df[original_df["accident"].astype(str) == "0"].copy()
        b = original_df[original_df["accident"].astype(str) == "1"].copy()
        if len(a) > 0 and len(b) > 0:
            return a.reset_index(drop=True), b.reset_index(drop=True), "Split by accident flag: 0 -> Dataset_A, 1 -> Dataset_B."

    half = max(1, len(original_df) // 2)
    a = original_df.iloc[:half].copy()
    b = original_df.iloc[half:].copy()
    return a.reset_index(drop=True), b.reset_index(drop=True), "Split by row halves due to missing/imbalanced accident flag."


def fuse_datasets(datasets: dict[str, pd.DataFrame]) -> tuple[pd.DataFrame, list[str]]:
    normalized: list[pd.DataFrame] = []
    all_cols: set[str] = set()
    for name, df in datasets.items():
        tmp = df.copy()
        tmp["source_dataset"] = name
        normalized.append(tmp)
        all_cols.update(tmp.columns.tolist())

    ordered = [c for c in BASE_COLS if c in all_cols] + [c for c in sorted(all_cols) if c not in BASE_COLS]
    aligned = [d.reindex(columns=ordered) for d in normalized]
    fused = pd.concat(aligned, ignore_index=True)
    pre = len(fused)
    fused = fused.drop_duplicates().reset_index(drop=True)
    removed = pre - len(fused)

    explanation = [
        "Fusion method: schema-aligned concatenation with source tracking.",
        "Why chosen: source datasets may not share reliable unique keys; row-wise append preserves all available observations.",
        "Step 1: add `source_dataset` column to each dataset.",
        "Step 2: align columns via union schema and reorder with core safety columns first.",
        "Step 3: concatenate datasets vertically.",
        f"Step 4: remove exact duplicate rows to avoid duplicate records (removed={removed}).",
    ]
    return fused, explanation


def _comparison_table(original: pd.DataFrame, image_df: pd.DataFrame, fused: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for name, df in [("Original", original), ("Image", image_df), ("Fused", fused)]:
        ndf = df.select_dtypes(include=[np.number])
        rows.append(
            {
                "dataset": name,
                "row_count": len(df),
                "numeric_columns": ndf.shape[1],
                "global_mean": float(ndf.mean(numeric_only=True).mean()) if not ndf.empty else np.nan,
                "global_std": float(ndf.std(numeric_only=True).mean()) if not ndf.empty else np.nan,
            }
        )
    return pd.DataFrame(rows)


def _combined_charts(fused_df: pd.DataFrame, out_dir: Path) -> tuple[list[Path], list[str]]:
    _require_plotting()
    out_dir.mkdir(parents=True, exist_ok=True)
    out: list[Path] = []
    notes: list[str] = []
    ndf = fused_df.select_dtypes(include=[np.number]).copy()
    if ndf.empty:
        return out, ["Combined analysis charts skipped: no numeric data in fused dataset."]

    cols = _safe_numeric_cols(ndf)
    if "source_dataset" in fused_df.columns and len(cols) >= 1:
        grp = fused_df.groupby("source_dataset")[cols[:3]].mean(numeric_only=True)
        fig, ax = plt.subplots(figsize=(10, 5))
        for c in grp.columns:
            ax.plot(grp.index.astype(str), grp[c], marker="o", label=c)
        ax.set_title("Comparative Line Chart - Mean by Source Dataset")
        ax.set_ylabel("Mean Value")
        ax.legend()
        p = out_dir / "fused_comparative_line.png"
        fig.tight_layout()
        fig.savefig(p, dpi=130)
        plt.close(fig)
        out.append(p)
        notes.append("Comparative line chart: contrasts average feature levels across source datasets and highlights cross-source trend shifts.")

        fig, ax = plt.subplots(figsize=(10, 5))
        grp.plot(kind="bar", ax=ax)
        ax.set_title("Multi-variable Bar Chart - Source vs Features")
        ax.set_ylabel("Mean Value")
        ax.tick_params(axis="x", rotation=0)
        p = out_dir / "fused_multivariable_bar.png"
        fig.tight_layout()
        fig.savefig(p, dpi=130)
        plt.close(fig)
        out.append(p)
        notes.append("Multi-variable bar chart: enables side-by-side comparison of multiple features per dataset source.")

    corr = ndf[cols].corr(numeric_only=True)
    fig, ax = plt.subplots(figsize=(8, 6))
    if sns is not None:
        sns.heatmap(corr, annot=True, fmt=".2f", cmap="coolwarm", ax=ax, vmin=-1, vmax=1)
    else:
        im = ax.imshow(corr.values, cmap="coolwarm", vmin=-1, vmax=1, aspect="auto")
        ax.set_xticks(range(len(corr.columns)))
        ax.set_xticklabels(corr.columns, rotation=45, ha="right")
        ax.set_yticks(range(len(corr.index)))
        ax.set_yticklabels(corr.index)
        fig.colorbar(im, ax=ax, label="Pearson r")
    ax.set_title("Correlation Heatmap - Fused Dataset")
    p = out_dir / "fused_correlation_heatmap.png"
    fig.tight_layout()
    fig.savefig(p, dpi=130)
    plt.close(fig)
    out.append(p)
    notes.append("Correlation heatmap: identifies positive/negative linear relationships and potential multicollinearity.")
    return out, notes


def _first_non_empty_sheet(workbook_path: Path) -> pd.DataFrame:
    sheets = pd.read_excel(workbook_path, sheet_name=None, engine="openpyxl")
    for _, df in sheets.items():
        if df is not None and not df.empty:
            return df.copy()
    raise ValueError("No non-empty sheets found in input workbook.")


def build_extension_report(input_excel: Path, input_image: Path, output_excel: Path) -> Path:
    original_raw = _first_non_empty_sheet(input_excel)
    image_result = extract_image_data(input_image)

    original_clean = clean_for_analysis(original_raw)
    image_clean = clean_for_analysis(image_result.mapped_table)
    dataset_a_raw, dataset_b_raw, split_reason = split_datasets(original_raw)
    dataset_a_clean = clean_for_analysis(dataset_a_raw)
    dataset_b_clean = clean_for_analysis(dataset_b_raw)

    fused_raw, fusion_explain = fuse_datasets(
        {
            "Original": original_clean,
            "Image": image_clean,
            "Dataset_A": dataset_a_clean,
            "Dataset_B": dataset_b_clean,
        }
    )
    fused_clean = clean_for_analysis(fused_raw)

    charts_dir = output_excel.parent / "_fusion_charts"
    if charts_dir.exists():
        shutil.rmtree(charts_dir)
    charts_dir.mkdir(parents=True, exist_ok=True)

    orig_charts, orig_notes = generate_individual_charts(original_clean, "Original", charts_dir)
    img_charts, img_notes = generate_individual_charts(image_clean, "Image", charts_dir)
    a_charts, a_notes = generate_individual_charts(dataset_a_clean, "Dataset_A", charts_dir)
    b_charts, b_notes = generate_individual_charts(dataset_b_clean, "Dataset_B", charts_dir)
    fused_charts, fused_notes = _combined_charts(fused_clean, charts_dir)

    wb = Workbook()

    # Sheet 1
    ws1 = wb.active
    ws1.title = "Original Data"
    _write_df(ws1, original_raw, title="Original dataset copy (unchanged).")
    _autofit_cols(ws1)

    # Sheet 2
    ws2 = wb.create_sheet("Image Extracted Data")
    r = _write_df(ws2, image_result.raw_table, title="Raw OCR/vision extraction from image.")
    r = _write_df(ws2, image_result.mapped_table, start_row=r, title="Mapped image data (meaningful columns).")
    ws2.cell(row=r, column=1, value="Detection Summary").font = Font(bold=True)
    ws2.cell(row=r + 1, column=1, value=f"Detected content: {image_result.detected_content}")
    ws2.cell(row=r + 2, column=1, value=f"Extraction method: {image_result.extraction_method}")
    ws2.cell(row=r + 3, column=1, value="Assumptions:")
    for i, txt in enumerate(image_result.assumptions[:12], start=r + 4):
        ws2.cell(row=i, column=1, value=f"- {txt}")
    k = r + 4 + len(image_result.assumptions[:12]) + 1
    ws2.cell(row=k, column=1, value="Corrections:")
    for i, txt in enumerate(image_result.corrections[:12], start=k + 1):
        ws2.cell(row=i, column=1, value=f"- {txt}")
    _autofit_cols(ws2)

    # Sheet 3
    ws3 = wb.create_sheet("Dataset_A")
    r3 = _write_df(ws3, dataset_a_raw, title="Dataset_A (split dataset)")
    r3 = _write_df(ws3, stats_table(dataset_a_clean, "Dataset_A"), start_row=r3, title="Dataset_A Statistical Summary")
    ws3.cell(row=r3, column=1, value=f"Split logic: {split_reason}")
    for i, note in enumerate(a_notes, start=r3 + 1):
        ws3.cell(row=i, column=1, value=f"- {note}")
    _embed_images(ws3, a_charts, "A40")
    _autofit_cols(ws3)

    # Sheet 4
    ws4 = wb.create_sheet("Dataset_B")
    r4 = _write_df(ws4, dataset_b_raw, title="Dataset_B (split dataset)")
    r4 = _write_df(ws4, stats_table(dataset_b_clean, "Dataset_B"), start_row=r4, title="Dataset_B Statistical Summary")
    ws4.cell(row=r4, column=1, value=f"Split logic: {split_reason}")
    for i, note in enumerate(b_notes, start=r4 + 1):
        ws4.cell(row=i, column=1, value=f"- {note}")
    _embed_images(ws4, b_charts, "A40")
    _autofit_cols(ws4)

    # Sheet 5
    ws5 = wb.create_sheet("Individual Analysis - Original")
    r5 = _write_df(ws5, stats_table(original_clean, "Original"), title="Original Dataset Statistical Analysis")
    for i, note in enumerate(orig_notes, start=r5):
        ws5.cell(row=i, column=1, value=f"- {note}")
    _embed_images(ws5, orig_charts, "A30")
    _autofit_cols(ws5)

    # Sheet 6
    ws6 = wb.create_sheet("Individual Analysis - Image Data")
    r6 = _write_df(ws6, stats_table(image_clean, "Image"), title="Image Dataset Statistical Analysis")
    for i, note in enumerate(img_notes, start=r6):
        ws6.cell(row=i, column=1, value=f"- {note}")
    _embed_images(ws6, img_charts, "A30")
    _autofit_cols(ws6)

    # Sheet 7
    ws7 = wb.create_sheet("Fused Dataset")
    r7 = _write_df(ws7, fused_raw, title="Fused dataset (new; source-tracked).")
    ws7.cell(row=r7, column=1, value="Fusion Explanation").font = Font(bold=True)
    for i, step in enumerate(fusion_explain, start=r7 + 1):
        ws7.cell(row=i, column=1, value=f"- {step}")
    _autofit_cols(ws7)

    # Sheet 8
    ws8 = wb.create_sheet("Combined Analysis")
    r8 = _write_df(ws8, stats_table(fused_clean, "Fused"), title="Fused Dataset Statistical Analysis")
    for i, note in enumerate(fused_notes, start=r8):
        ws8.cell(row=i, column=1, value=f"- {note}")
    _embed_images(ws8, fused_charts, "A30")
    _autofit_cols(ws8)

    # Sheet 9
    ws9 = wb.create_sheet("Final Comparison")
    comp = _comparison_table(original_clean, image_clean, fused_clean)
    r9 = _write_df(ws9, comp, title="Comparison: Original vs Image vs Fused")
    insights = [
        "Image-based extraction introduces additional observations that may not exist in historical logs.",
        "Fusion improves data diversity and can reveal trends hidden in single-source analysis.",
        "Any divergence between original and image distributions should be treated as anomaly candidates for manual review.",
        "Use fused trends to prioritize feature engineering and risk control interventions.",
    ]
    ws9.cell(row=r9, column=1, value="Final Insights (Academic Tone)").font = Font(bold=True)
    for i, t in enumerate(insights, start=r9 + 1):
        ws9.cell(row=i, column=1, value=f"- {t}")
    _autofit_cols(ws9)

    output_excel.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_excel)
    return output_excel


def main() -> None:
    parser = argparse.ArgumentParser(description="AccidentZero extension: image extraction + data fusion + Excel analysis.")
    parser.add_argument("--input-excel", required=True, help="Path to existing workbook (original project output).")
    parser.add_argument("--input-image", required=True, help="Path to source image containing table/chart data.")
    parser.add_argument("--output-excel", required=True, help="Path for NEW fused analysis workbook.")
    args = parser.parse_args()

    out = build_extension_report(
        input_excel=Path(args.input_excel).expanduser().resolve(),
        input_image=Path(args.input_image).expanduser().resolve(),
        output_excel=Path(args.output_excel).expanduser().resolve(),
    )
    print(f"[OK] Extension workbook generated: {out}")


if __name__ == "__main__":
    main()
